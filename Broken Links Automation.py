#!/usr/bin/env python
__author__ = "Ashutosh Mishra"
__credits__ = ["Ashutosh Mishra"]
__code_name__ = "Broken Links Automation"
__version__ = "1.0"
__maintainer__ = "Ashutosh Mishra"
__status__ = "Production"

# importing libraries #

import requests

import os

import time

from openpyxl import load_workbook

import pandas as pd

from datetime import datetime

 
# getting directory path #

dir_path = os.path.dirname(os.path.realpath(__file__))

 
# main class #

class Checker:

 # initializing empty lists and dictionaries #
 
    def __init__(self):

        self.input = []

        self.output = {'Link': [], 'Status': [],'Redirected_Links':[]}

 
# reading input links from input file #

    def read_input(self,link_type):

        input_file = dir_path + '\Input\input_links.xlsx'

        wb = load_workbook(input_file)

        ws = wb[link_type]

        max_row = ws.max_row

        max_col = ws.max_column

        for row in ws.iter_rows(min_row=0, min_col=0, max_row= max_row, max_col= max_col):

            for cell in row:

                if cell.value:

                    self.input.append(cell.value)

        wb.close()

        print('Reading',link_type,'Done')

# main function code that determines links are broken or working #
 
    def crawler(self):

        count = 1

        for link in self.input:

            time.sleep(1)

            print(f'Processing links: {count} of {len(self.input)}')

            self.output['Link'].append(link)

            try:

                source = str(requests.get(link))

                src_url = requests.get(link)

                if source == '<Response [404]>':

                    self.output['Status'].append('Not Found (404)')

                    self.output['Redirected_Links'].append(src_url.url)

                else:

                    self.output['Status'].append('Working')

                    self.output['Redirected_Links'].append(src_url.url)

            except:

                self.output['Status'].append('Error')

                self.output['Redirected_Links'].append('Error')

            count += 1

 
# write output file fuction #

    def write_output(self,output_sheet,filename):

        df = pd.DataFrame(self.output)

        now = datetime.now().strftime("_%d_%b_%y_%I_%M_%p")

        writer = pd.ExcelWriter(dir_path + '\Output\Output-'+filename + str(now) + '.xlsx')

        df.to_excel(writer,output_sheet, index=False)

        writer.save()

        self.input.clear()

        for key in self.output:

            self.output[key].clear()

 
# main program starts from here #

obj = Checker()

t1 = time.time()

obj.read_input('Parent_Links')

obj.crawler()

obj.write_output('Parent_Links Status','Parent')

print('')

obj.read_input('Child_Links')

obj.crawler()

obj.write_output('Child_Links Status','Child')

t2 = time.time()

print(f'Completed,Total time taken: {(t2-t1)/60} mins')
 
